from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import List, Dict, Optional, Tuple

from openpyxl import load_workbook


@dataclass
class BudgetItem:
    category: str
    amount_gbp: float
    amount_ngn: Optional[float] = None
    notes: str = ""


@dataclass
class CarePlan:
    weekly_income_gbp: float
    monthly_income_gbp: float
    total_support_gbp: float
    total_support_ngn: Optional[float]
    items: List[BudgetItem]


def _to_float(v) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    # Handle strings like "£200"
    for ch in ["£", "GBP", "NGN", "₦"]:
        s = s.replace(ch, "").strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _to_ngn_float(v) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    # Handle strings like "₦216,000" or "NGN 216,000"
    s = s.replace("₦", "").replace("NGN", "").strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def read_care_plan_from_excel(xlsx_path: str | Path) -> CarePlan:
    """
    Reads the 'Mom Monthly Support Plan' sheet created earlier.
    Expected layout:
      - Column A: labels/categories
      - Column B: GBP numeric values for incomes + support amounts
      - Column C: NGN values (some are strings like ₦216,000)
      - Column D: notes
    """
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Excel file not found: {xlsx_path}")

    wb = load_workbook(xlsx_path, data_only=True)
    sheet_name = "Mom Monthly Support Plan"
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Found: {wb.sheetnames}")

    ws = wb[sheet_name]

    # Helper: find row by label in col A
    def find_row(label: str) -> Optional[int]:
        label_low = label.strip().lower()
        for r in range(1, ws.max_row + 1):
            a = ws.cell(row=r, column=1).value
            if a is None:
                continue
            if str(a).strip().lower() == label_low:
                return r
        return None

    # Income rows
    weekly_row = find_row("Weekly Income")
    monthly_row = find_row("Monthly Income")

    if weekly_row is None or monthly_row is None:
        raise ValueError("Could not find 'Weekly Income' and/or 'Monthly Income' rows in column A.")

    weekly_income = _to_float(ws.cell(row=weekly_row, column=2).value) or 0.0
    monthly_income = _to_float(ws.cell(row=monthly_row, column=2).value) or 0.0

    # Budget items are under the header "MONTHLY SUPPORT BREAKDOWN"
    header_row = find_row("MONTHLY SUPPORT BREAKDOWN")
    if header_row is None:
        raise ValueError("Could not find 'MONTHLY SUPPORT BREAKDOWN' header row.")

    items: List[BudgetItem] = []
    total_support_gbp: float = 0.0
    total_support_ngn: Optional[float] = None

    # Read rows after the header until we hit TOTAL MONTHLY SUPPORT (or blank block ends)
    r = header_row + 1
    while r <= ws.max_row:
        cat = ws.cell(row=r, column=1).value
        gbp = ws.cell(row=r, column=2).value
        ngn = ws.cell(row=r, column=3).value
        note = ws.cell(row=r, column=4).value

        if cat is None:
            r += 1
            continue

        cat_str = str(cat).strip()
        if not cat_str:
            r += 1
            continue

        # Stop when we reach totals
        if cat_str.upper() in {"TOTAL MONTHLY SUPPORT", "REMAINING FOR YOU"}:
            if cat_str.upper() == "TOTAL MONTHLY SUPPORT":
                total_support_gbp = _to_float(gbp) or 0.0
                total_support_ngn = _to_ngn_float(ngn)
            break

        # Skip empty separators
        if cat_str == "":
            r += 1
            continue

        amount_gbp = _to_float(gbp)
        # If we are in breakdown region but row has no GBP, skip
        if amount_gbp is None:
            r += 1
            continue

        amount_ngn = _to_ngn_float(ngn)
        items.append(
            BudgetItem(
                category=cat_str,
                amount_gbp=amount_gbp,
                amount_ngn=amount_ngn,
                notes=str(note).strip() if note else "",
            )
        )
        r += 1

    # If totals weren't found, compute from items
    if total_support_gbp == 0.0 and items:
        total_support_gbp = sum(i.amount_gbp for i in items)

    return CarePlan(
        weekly_income_gbp=weekly_income,
        monthly_income_gbp=monthly_income,
        total_support_gbp=total_support_gbp,
        total_support_ngn=total_support_ngn,
        items=items,
    )


def format_plan_for_telegram(plan: CarePlan) -> str:
    """
    Returns a clean Telegram-friendly message.
    """
    lines = []
    lines.append("🧾 MomCareBot — Monthly Support Plan")
    lines.append(f"Income: £{plan.weekly_income_gbp:.0f}/week (~£{plan.monthly_income_gbp:.0f}/month)")
    if plan.total_support_ngn is not None:
        lines.append(f"Planned support: £{plan.total_support_gbp:.0f} (≈ ₦{plan.total_support_ngn:,.0f})")
    else:
        lines.append(f"Planned support: £{plan.total_support_gbp:.0f}")

    lines.append("")
    lines.append("Breakdown:")
    for item in plan.items:
        if item.amount_ngn is not None:
            lines.append(f"• {item.category}: £{item.amount_gbp:.0f} (≈ ₦{item.amount_ngn:,.0f})")
        else:
            lines.append(f"• {item.category}: £{item.amount_gbp:.0f}")
    lines.append("")
    lines.append("✅ Reminder: keep it consistent + save emergency fund monthly.")
    return "\n".join(lines)
